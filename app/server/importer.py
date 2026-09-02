from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from re import sub
from uuid import uuid4

from openpyxl import load_workbook
from pypdf import PdfReader

from .models import ImportSummary, ParticipantGroup, ProductLine, TrainingProject, TrainingTopic
from .planner import plan_project

DEFAULT_PDF_TOPICS = {
    "PACS-Administration": 360,
    "DU Diagnost Basic": 90,
    "DU Diagnost erweitert": 120,
    "DU Diagnost KeyUser": 180,
    "DU Review Kliniker": 60,
    "DU Viewer": 45,
    "DU XChange": 60,
    "DU Review MTRA": 60,
}


def _topic_id(title: str) -> str:
    clean = sub(r"[^a-zA-Z0-9]+", "-", title.strip().lower()).strip("-")
    return clean or uuid4().hex[:8]


def analyze_excel(name: str, data: bytes) -> tuple[dict, TrainingProject]:
    workbook = load_workbook(BytesIO(data), data_only=True, keep_vba=True)
    sheets: list[dict] = []
    project = TrainingProject()
    seen_topics: dict[str, TrainingTopic] = {}

    for worksheet in workbook.worksheets:
        non_empty_rows = []
        for row in worksheet.iter_rows():
            values = [cell.value for cell in row]
            if any(value not in (None, "") for value in values):
                non_empty_rows.append(values)

        sheet_summary = {
            "name": worksheet.title,
            "rows": worksheet.max_row,
            "columns": worksheet.max_column,
            "nonEmptyRows": len(non_empty_rows),
            "headers": [],
            "sampleRows": [],
        }

        headers = [cell.value for cell in worksheet[12] if cell.value not in (None, "")]
        sheet_summary["headers"] = [str(header).replace("\n", " ") for header in headers]
        for values in non_empty_rows[:8]:
            sheet_summary["sampleRows"].append([str(value) for value in values if value not in (None, "")][:8])

        if worksheet["K6"].value:
            project.customer_name = str(worksheet["K6"].value)
            project.location = str(worksheet["K6"].value)
        if worksheet["K8"].value:
            project.trainer = str(worksheet["K8"].value)
        if worksheet["K10"].value and project.start_date is None:
            if isinstance(worksheet["K10"].value, datetime):
                project.start_date = worksheet["K10"].value.date()
            else:
                try:
                    project.start_date = datetime.fromisoformat(str(worksheet["K10"].value)).date()
                except ValueError:
                    pass

        group_rows = [(47, "Diagnost"), (48, "Review"), (49, "MTRA")]
        groups = []
        for row_index, label in group_rows:
            participant_count = worksheet[f"J{row_index}"].value
            sessions = worksheet[f"K{row_index}"].value
            if participant_count:
                group_id = _topic_id(f"deepunity-pacs-{label}")
                groups.append(ParticipantGroup(
                    id=group_id,
                    product_id="deepunity-pacs",
                    name=label,
                    participant_count=int(float(participant_count)),
                    notes=f"{sessions or 1} Schulungsbloecke laut Excel.",
                ))
                title = f"{label} Schulungsgruppe"
                seen_topics[_topic_id(title)] = TrainingTopic(
                    id=_topic_id(title),
                    product_id="deepunity-pacs",
                    participant_group_id=group_id,
                    title=title,
                    description=f"Aus Excel-Blatt {worksheet.title} abgeleitete Gruppe.",
                    duration_minutes=max(45, int(float(sessions or 1) * 45)),
                    priority=3,
                    participants_per_session=int(float(participant_count)),
                    sessions_required=float(sessions or 1),
                )

        if groups:
            project.product_lines = [ProductLine(
                id="deepunity-pacs",
                name="DeepUnity PACS",
                description="PACS-Schulungen fuer Radiologie, Keyuser, MFA, Kliniker, Webviewer und Administration.",
                participant_groups=groups,
            )]

        sheets.append(sheet_summary)

    project.title = "DeepUnity Schulungsplan"
    project.participant_group = "Diagnost / Review / MTRA"
    project.end_date = None
    project.topics = list(seen_topics.values())
    return {"fileName": name, "sheets": sheets}, project


def analyze_pdf(name: str, data: bytes) -> dict:
    reader = PdfReader(BytesIO(data))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    title = Path(name).stem
    duration = DEFAULT_PDF_TOPICS.get(title, 60)
    for index, line in enumerate(lines):
        if line == "Wie lang?" and index + 1 < len(lines):
            next_line = lines[index + 1]
            if "0.75" in next_line or "0,75" in next_line:
                duration = 45
            elif "1,5" in next_line or "1.5" in next_line:
                duration = 90
            elif "2" in next_line:
                duration = 120
            elif "3" in next_line:
                duration = 180
            elif "6" in next_line:
                duration = 360
    topics = []
    for index, line in enumerate(lines):
        if "min" in line and index + 1 < len(lines):
            topics.append(lines[index + 1])
    return {
        "fileName": name,
        "title": title,
        "pages": len(reader.pages),
        "estimatedDurationMinutes": duration,
        "structure": ["Allgemeine Informationen", "Vorbereitung", "Schulungsverlauf"],
        "sampleTopics": list(dict.fromkeys(topics[:10])),
        "textPreview": lines[:18],
    }


def build_project_from_uploads(excel_files: list[tuple[str, bytes]], pdf_files: list[tuple[str, bytes]]) -> ImportSummary:
    excel_summaries: list[dict] = []
    project = TrainingProject()
    for name, data in excel_files:
        summary, imported_project = analyze_excel(name, data)
        excel_summaries.append(summary)
        project.location = imported_project.location or project.location
        project.customer_name = imported_project.customer_name or project.customer_name
        project.trainer = imported_project.trainer or project.trainer
        project.participant_group = imported_project.participant_group or project.participant_group
        project.start_date = imported_project.start_date or project.start_date
        project.product_lines = imported_project.product_lines or project.product_lines
        project.topics.extend(imported_project.topics)

    pdf_summaries = [analyze_pdf(name, data) for name, data in pdf_files]
    existing_titles = {topic.title for topic in project.topics}
    for pdf in pdf_summaries:
        if pdf["title"] not in existing_titles:
            project.topics.append(TrainingTopic(
                id=_topic_id(pdf["title"]),
                product_id="deepunity-pacs",
                title=pdf["title"],
                description="Aus PDF-Vorlage abgeleiteter Schulungsbaustein.",
                duration_minutes=pdf["estimatedDurationMinutes"],
                priority=2,
                notes=", ".join(pdf["sampleTopics"][:3]),
            ))

    project = plan_project(project)
    return ImportSummary(excel={"files": excel_summaries}, pdfs=pdf_summaries, project=project)
